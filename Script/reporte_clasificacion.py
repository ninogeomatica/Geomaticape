"""
Reporte de Clasificacion Raster
================================
Genera un reporte estadistico completo a partir de un raster clasificado
(valores categoricos enteros). Salida: Excel (.xlsx) con hoja de resumen
y graficos, mas CSV complementario.

Autor : Geomatica Ambiental - https://www.geomatica.pe
Plugin: Geomaticape v1.2
Grupo : PostProcesamiento
"""

import os
import gc
import math
import csv
import numpy as np

from qgis.PyQt.QtGui import QIcon
from qgis.core import (
    QgsProcessingAlgorithm,
    QgsProcessingParameterRasterLayer,
    QgsProcessingParameterBand,
    QgsProcessingParameterMatrix,
    QgsProcessingParameterFileDestination,
    QgsProcessingException,
)
from qgis import processing
from osgeo import gdal


class ReporteClasificacion(QgsProcessingAlgorithm):

    INPUT_RASTER = "INPUT_RASTER"
    BAND         = "BAND"
    LEGEND_TABLE = "LEGEND_TABLE"
    OUTPUT       = "OUTPUT"

    # -------------------------------------------------------
    # IDENTIFICACION
    # -------------------------------------------------------

    def name(self):
        return "reporte_clasificacion"

    def displayName(self):
        return "Reporte de clasificacion raster (area, porcentaje, estadisticas)"

    def group(self):
        return "PostProcesamiento"

    def groupId(self):
        return "geomaticape_postprocesamiento"

    def icon(self):
        return QIcon(os.path.join(os.path.dirname(__file__), "..", "Icons", "zonal_raster.png"))

    def createInstance(self):
        return ReporteClasificacion()

    # -------------------------------------------------------
    # AYUDA
    # -------------------------------------------------------

    def shortHelpString(self):
        return """
<h3>Reporte de Clasificacion Raster</h3>
<b>Autor:</b> GEOMATICA AMBIENTAL<br>
<b>Plugin:</b> Geomaticape<br>
<b>Version:</b> 1.2<br><br>

<b>Descripcion:</b><br>
Genera un reporte estadistico completo de un raster clasificado.
Para cada clase calcula metricas de area, cobertura y forma.<br><br>

<b>Metricas calculadas por clase:</b>
<ul>
<li><b>N pixeles</b>: conteo total de pixeles de la clase</li>
<li><b>Area (m2)</b>: pixeles x resolucion al cuadrado</li>
<li><b>Area (ha)</b>: area en hectareas</li>
<li><b>Area (km2)</b>: area en kilometros cuadrados</li>
<li><b>% sobre area total</b>: incluyendo NoData</li>
<li><b>% sobre area valida</b>: excluyendo NoData</li>
<li><b>Pixeles frontera</b>: estimacion del perimetro interno de la clase</li>
<li><b>Perimetro (m)</b>: longitud estimada del perimetro en metros</li>
<li><b>Indice de forma</b>: perimetro / (2 x sqrt(pi x area)), mayor valor indica mayor irregularidad</li>
</ul>

<b>Leyenda de clases (opcional):</b><br>
Ingresa en la tabla: <b>Valor clase | Etiqueta</b><br>
Si se deja vacia se usa "Clase N".<br><br>

<b>Salida:</b><br>
Excel (.xlsx) con:
<ul>
<li><b>Hoja Resumen</b>: tabla completa con metadatos del raster y metricas por clase</li>
<li><b>Hoja Graficos</b>: grafico de pastel (% area valida) y grafico de barras (ha)</li>
</ul>
Ademas se genera un CSV con la misma tabla para uso en SIG o scripts.<br><br>

<b>Web:</b> https://www.geomatica.pe/
"""

    # -------------------------------------------------------
    # PARAMETROS
    # -------------------------------------------------------

    def initAlgorithm(self, config=None):

        self.addParameter(
            QgsProcessingParameterRasterLayer(
                self.INPUT_RASTER,
                "Raster clasificado (valores categoricos enteros)"
            )
        )

        self.addParameter(
            QgsProcessingParameterBand(
                self.BAND,
                "Banda del raster clasificado",
                parentLayerParameterName=self.INPUT_RASTER,
                optional=False
            )
        )

        self.addParameter(
            QgsProcessingParameterMatrix(
                self.LEGEND_TABLE,
                "Leyenda de clases  (Valor clase | Etiqueta) - opcional",
                headers=["Valor clase", "Etiqueta"],
                defaultValue=[
                    1, "Agua / Nube",
                    2, "Suelo desnudo",
                    3, "Vegetacion escasa",
                    4, "Vegetacion moderada",
                    5, "Vegetacion densa",
                ],
                optional=True
            )
        )

        self.addParameter(
            QgsProcessingParameterFileDestination(
                self.OUTPUT,
                "Reporte Excel de salida",
                fileFilter="Excel (*.xlsx)"
            )
        )

    # -------------------------------------------------------
    # UTILIDADES
    # -------------------------------------------------------

    @staticmethod
    def _parse_legend(flat_list):
        """Parsea la tabla plana de leyenda y devuelve {valor_int: etiqueta}."""
        leyenda = {}
        if not flat_list or len(flat_list) < 2:
            return leyenda
        if len(flat_list) % 2 != 0:
            return leyenda
        for i in range(0, len(flat_list), 2):
            try:
                val   = int(float(str(flat_list[i])))
                label = str(flat_list[i + 1]).strip()
                leyenda[val] = label
            except (TypeError, ValueError):
                pass
        return leyenda

    @staticmethod
    def _border_pixels(mask):
        """
        Cuenta los pixeles en el borde interno de una mascara binaria
        (4-conectividad). Un pixel es borde si es True y tiene al menos
        un vecino False. Implementacion por desplazamiento de arrays del
        mismo shape — sin riesgo de mismatch de dimensiones.
        """
        m = mask.astype(bool)

        up    = np.zeros_like(m)
        down  = np.zeros_like(m)
        left  = np.zeros_like(m)
        right = np.zeros_like(m)

        up[1:, :]     = m[:-1, :]   # vecino de arriba
        down[:-1, :]  = m[1:,  :]   # vecino de abajo
        left[:, 1:]   = m[:, :-1]   # vecino izquierdo
        right[:, :-1] = m[:, 1:]    # vecino derecho

        border = m & (~up | ~down | ~left | ~right)
        return int(np.sum(border))

    # -------------------------------------------------------
    # PROCESO PRINCIPAL
    # -------------------------------------------------------

    def processAlgorithm(self, parameters, context, feedback):

        # ---------------------------------------------------
        # 1. PARAMETROS
        # ---------------------------------------------------
        raster_layer = self.parameterAsRasterLayer(parameters, self.INPUT_RASTER, context)
        if raster_layer is None:
            raise QgsProcessingException("No se pudo leer el raster clasificado.")

        input_path  = raster_layer.source()
        band_num    = self.parameterAsInt(parameters, self.BAND, context)
        legend_flat = self.parameterAsMatrix(parameters, self.LEGEND_TABLE, context)
        output_path = self.parameterAsFileOutput(parameters, self.OUTPUT, context)

        # Asegurar extension .xlsx
        if not output_path.lower().endswith(".xlsx"):
            output_path = os.path.splitext(output_path)[0] + ".xlsx"

        # CSV paralelo en la misma carpeta
        csv_path = os.path.splitext(output_path)[0] + "_reporte.csv"

        leyenda = self._parse_legend(legend_flat)

        feedback.pushInfo("=" * 52)
        feedback.pushInfo("Raster clasificado : {}".format(os.path.basename(input_path)))
        feedback.pushInfo("Banda              : {}".format(band_num))
        feedback.pushInfo("Clases en leyenda  : {}".format(len(leyenda)))
        feedback.pushInfo("Salida Excel       : {}".format(os.path.basename(output_path)))
        feedback.pushInfo("Salida CSV         : {}".format(os.path.basename(csv_path)))
        feedback.pushInfo("=" * 52)

        # ---------------------------------------------------
        # 2. LECTURA DEL RASTER CLASIFICADO
        # ---------------------------------------------------
        ds = gdal.Open(input_path, gdal.GA_ReadOnly)
        if ds is None:
            raise QgsProcessingException("GDAL no pudo abrir el raster clasificado.")

        n_cols = ds.RasterXSize
        n_rows = ds.RasterYSize
        gt     = ds.GetGeoTransform()
        nbands = ds.RasterCount

        if band_num < 1 or band_num > nbands:
            raise QgsProcessingException(
                "Banda {} no existe (el raster tiene {} banda(s)).".format(band_num, nbands)
            )

        pixel_width   = abs(gt[1])
        pixel_height  = abs(gt[5])
        pixel_area_m2 = pixel_width * pixel_height

        feedback.pushInfo("Dimensiones   : {} x {} px".format(n_cols, n_rows))
        feedback.pushInfo("Resolucion    : {:.4f} x {:.4f} m".format(pixel_width, pixel_height))
        feedback.pushInfo("Area pixel    : {:.4f} m2".format(pixel_area_m2))

        b   = ds.GetRasterBand(band_num)
        arr = b.ReadAsArray()
        nd  = b.GetNoDataValue()
        ds  = None

        if nd is not None:
            nd_cmp = int(round(nd)) if arr.dtype.kind in ('i', 'u') else nd
            nodata_mask = (arr == nd_cmp)
        else:
            nodata_mask = np.zeros(arr.shape, dtype=bool)

        feedback.setProgress(20)

        # ---------------------------------------------------
        # 3. CLASES UNICAS
        # ---------------------------------------------------
        feedback.pushInfo("Identificando clases...")
        valid_arr      = arr[~nodata_mask]
        clases_unicas  = sorted(np.unique(valid_arr).tolist())
        n_clases       = len(clases_unicas)
        n_px_total     = int(arr.size)
        n_px_validos   = int(valid_arr.size)
        n_px_nodata    = int(np.sum(nodata_mask))
        area_total_ha  = n_px_total  * pixel_area_m2 / 10_000
        area_valida_ha = n_px_validos * pixel_area_m2 / 10_000

        feedback.pushInfo("Clases encontradas : {}".format(n_clases))
        feedback.pushInfo("Px totales         : {:,}".format(n_px_total))
        feedback.pushInfo("Px validos         : {:,}".format(n_px_validos))
        feedback.pushInfo("Px NoData          : {:,}".format(n_px_nodata))
        feedback.pushInfo("Area valida total  : {:,.4f} ha".format(area_valida_ha))
        feedback.setProgress(35)

        # ---------------------------------------------------
        # 4. METRICAS POR CLASE
        # ---------------------------------------------------
        feedback.pushInfo("Calculando metricas por clase...")
        registros = []

        for i, cls in enumerate(clases_unicas):
            cls_int = int(cls)
            mask    = (arr == cls) & ~nodata_mask
            n_pix   = int(np.sum(mask))

            area_m2  = n_pix * pixel_area_m2
            area_ha  = area_m2 / 10_000
            area_km2 = area_m2 / 1_000_000
            pct_total  = (n_pix / n_px_total  * 100) if n_px_total  > 0 else 0.0
            pct_valido = (n_pix / n_px_validos * 100) if n_px_validos > 0 else 0.0

            n_border    = self._border_pixels(mask)
            perimetro_m = n_border * ((pixel_width + pixel_height) / 2.0)

            if area_m2 > 0:
                shape_idx = perimetro_m / (2.0 * math.sqrt(math.pi * area_m2))
            else:
                shape_idx = 0.0

            etiqueta = leyenda.get(cls_int, "Clase {}".format(cls_int))

            registros.append({
                "Valor_Clase":     cls_int,
                "Etiqueta":        etiqueta,
                "N_Pixeles":       n_pix,
                "Area_m2":         round(area_m2,  2),
                "Area_ha":         round(area_ha,  4),
                "Area_km2":        round(area_km2, 6),
                "Pct_Area_Total":  round(pct_total,  4),
                "Pct_Area_Valida": round(pct_valido, 4),
                "Px_Frontera":     n_border,
                "Perimetro_m":     round(perimetro_m, 2),
                "Indice_Forma":    round(shape_idx, 4),
            })

            feedback.pushInfo(
                "  Clase {:>4} | {:<25} | {:>10,} px | {:>12,.4f} ha | {:>7.3f} %".format(
                    cls_int, etiqueta, n_pix, area_ha, pct_valido
                )
            )
            feedback.setProgress(35 + int((i + 1) / n_clases * 45))

        # Fila TOTAL
        registros.append({
            "Valor_Clase":     "TOTAL",
            "Etiqueta":        "Area valida total",
            "N_Pixeles":       n_px_validos,
            "Area_m2":         round(area_valida_ha * 10_000, 2),
            "Area_ha":         round(area_valida_ha, 4),
            "Area_km2":        round(area_valida_ha / 100, 6),
            "Pct_Area_Total":  round(n_px_validos / n_px_total * 100, 4) if n_px_total > 0 else 0,
            "Pct_Area_Valida": 100.0,
            "Px_Frontera":     "",
            "Perimetro_m":     "",
            "Indice_Forma":    "",
        })

        cols_out = [
            "Valor_Clase", "Etiqueta",
            "N_Pixeles", "Area_m2", "Area_ha", "Area_km2",
            "Pct_Area_Total", "Pct_Area_Valida",
            "Px_Frontera", "Perimetro_m", "Indice_Forma",
        ]

        feedback.setProgress(82)

        # ---------------------------------------------------
        # 5. CSV
        # ---------------------------------------------------
        feedback.pushInfo("Escribiendo CSV: {}".format(os.path.basename(csv_path)))
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols_out, extrasaction="ignore")
            w.writeheader()
            for r in registros:
                w.writerow(r)

        # ---------------------------------------------------
        # 6. EXCEL
        # ---------------------------------------------------
        feedback.pushInfo("Generando Excel: {}".format(os.path.basename(output_path)))
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.chart.label import DataLabelList
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise QgsProcessingException(
                "Falta la libreria 'openpyxl'. Instalala con:\n"
                "  python -m pip install openpyxl"
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"

        # -- Estilos --
        hdr_fill   = PatternFill("solid", fgColor="1F4E79")
        total_fill = PatternFill("solid", fgColor="BDD7EE")
        hdr_font   = Font(bold=True, color="FFFFFF", size=11)
        total_font = Font(bold=True, size=10)
        body_font  = Font(size=10)
        center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_al    = Alignment(horizontal="left",   vertical="center")
        thin       = Side(border_style="thin", color="AAAAAA")
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)

        # -- Titulo --
        ws.merge_cells("A1:D1")
        ws["A1"] = "REPORTE DE CLASIFICACION RASTER"
        ws["A1"].font      = Font(bold=True, size=14, color="1F4E79")
        ws["A1"].alignment = center

        # -- Metadatos --
        meta = [
            ("Raster:",            os.path.basename(input_path)),
            ("Resolucion:",        "{:.4f} x {:.4f} m  (pixel = {:.4f} m2)".format(
                                    pixel_width, pixel_height, pixel_area_m2)),
            ("Dimensiones:",       "{} x {} px".format(n_cols, n_rows)),
            ("N clases:",          n_clases),
            ("Px validos:",        "{:,}".format(n_px_validos)),
            ("Px NoData:",         "{:,}".format(n_px_nodata)),
            ("Area valida total:", "{:,.4f} ha".format(area_valida_ha)),
            ("Area total:",        "{:,.4f} ha".format(area_total_ha)),
        ]
        for row_i, (k, v) in enumerate(meta, start=2):
            ws["A{}".format(row_i)] = k
            ws["B{}".format(row_i)] = v
            ws["A{}".format(row_i)].font = Font(bold=True, size=10)
            ws["B{}".format(row_i)].font = Font(size=10)

        ROW_START = len(meta) + 3   # fila donde arranca la tabla (2 + 8 meta + 1 espacio = 11)

        # -- Encabezados de tabla --
        header_labels = {
            "Valor_Clase":     "Valor\nClase",
            "Etiqueta":        "Etiqueta / Categoria",
            "N_Pixeles":       "N Pixeles",
            "Area_m2":         "Area (m2)",
            "Area_ha":         "Area (ha)",
            "Area_km2":        "Area (km2)",
            "Pct_Area_Total":  "% Area\nTotal",
            "Pct_Area_Valida": "% Area\nValida",
            "Px_Frontera":     "Px\nFrontera",
            "Perimetro_m":     "Perimetro\n(m)",
            "Indice_Forma":    "Indice\nForma",
        }
        for col_i, col_key in enumerate(cols_out, start=1):
            cell = ws.cell(row=ROW_START, column=col_i,
                           value=header_labels.get(col_key, col_key))
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = border

        # Anchos de columna
        ws.row_dimensions[ROW_START].height = 30
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 28
        for col_i in range(3, len(cols_out) + 1):
            ws.column_dimensions[get_column_letter(col_i)].width = 14

        # -- Filas de datos --
        for r_i, rec in enumerate(registros):
            row_excel = ROW_START + 1 + r_i
            is_total  = (rec["Valor_Clase"] == "TOTAL")
            for col_i, col_key in enumerate(cols_out, start=1):
                val  = rec.get(col_key, "")
                cell = ws.cell(row=row_excel, column=col_i, value=val)
                cell.border = border
                cell.font   = total_font if is_total else body_font
                if is_total:
                    cell.fill = total_fill
                cell.alignment = left_al if col_key == "Etiqueta" else center
                if not is_total and isinstance(val, float):
                    if "Pct" in col_key:
                        cell.number_format = "0.0000"
                    elif col_key == "Area_ha":
                        cell.number_format = "#,##0.0000"
                    elif col_key == "Area_m2":
                        cell.number_format = "#,##0.00"
                    elif col_key in ("Area_km2", "Indice_Forma", "Perimetro_m"):
                        cell.number_format = "#,##0.0000"

        ws.freeze_panes = ws.cell(row=ROW_START + 1, column=1)

        # ---------------------------------------------------
        # 7. HOJA GRAFICOS
        # ---------------------------------------------------
        wg = wb.create_sheet("Graficos")
        reg_graf = [r for r in registros if r["Valor_Clase"] != "TOTAL"]
        n_reg    = len(reg_graf)

        wg["A1"] = "Etiqueta"
        wg["B1"] = "Area_ha"
        wg["C1"] = "Pct_Area_Valida"
        for gi, r in enumerate(reg_graf, start=2):
            wg.cell(row=gi, column=1, value=r["Etiqueta"])
            wg.cell(row=gi, column=2, value=r["Area_ha"])
            wg.cell(row=gi, column=3, value=r["Pct_Area_Valida"])

        # Grafico pastel (% area valida)
        pie = PieChart()
        pie.title  = "Distribucion de Area por Clase (%)"
        pie.style  = 10
        pie.width  = 18
        pie.height = 14
        data_pie = Reference(wg, min_col=3, min_row=1, max_row=1 + n_reg)
        cats_pie = Reference(wg, min_col=1, min_row=2, max_row=1 + n_reg)
        pie.add_data(data_pie, titles_from_data=True)
        pie.set_categories(cats_pie)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        wg.add_chart(pie, "E2")

        # Grafico barras (ha)
        bar = BarChart()
        bar.type   = "col"
        bar.title  = "Area por Clase (ha)"
        bar.style  = 10
        bar.y_axis.title = "Area (ha)"
        bar.x_axis.title = "Clase"
        bar.width  = 22
        bar.height = 14
        data_bar = Reference(wg, min_col=2, min_row=1, max_row=1 + n_reg)
        cats_bar = Reference(wg, min_col=1, min_row=2, max_row=1 + n_reg)
        bar.add_data(data_bar, titles_from_data=True)
        bar.set_categories(cats_bar)
        wg.add_chart(bar, "E20")

        wb.save(output_path)
        feedback.pushInfo("Excel guardado: {}".format(output_path))

        # ---------------------------------------------------
        # 8. RESUMEN FINAL
        # ---------------------------------------------------
        del arr
        gc.collect()

        feedback.setProgress(100)
        feedback.pushInfo("=" * 52)
        feedback.pushInfo("REPORTE DE CLASIFICACION COMPLETO")
        feedback.pushInfo("Clases          : {}".format(n_clases))
        feedback.pushInfo("Area valida     : {:,.4f} ha".format(area_valida_ha))
        feedback.pushInfo("Px NoData       : {:,}".format(n_px_nodata))
        feedback.pushInfo("Excel           : {}".format(output_path))
        feedback.pushInfo("CSV             : {}".format(csv_path))
        feedback.pushInfo("=" * 52)

        return {self.OUTPUT: output_path}

    def run(self):
        processing.execAlgorithmDialog(self)
